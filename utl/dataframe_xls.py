import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

from utl import add_timestamp_to_filename 

class Dataframe2XLS:
    def write_xls(self, df, excel_filename):
        excel_filename = add_timestamp_to_filename(excel_filename)
  
        wb = Workbook()
        ws = wb.active
        ws.title = "Daten"

        # 3️⃣ Kopfzeile schreiben (stellt sicher, dass alle Spalten korrekt erfasst werden)
        ws.append(df.columns.tolist())

        # 4️⃣ Daten einfügen (Führende Nullen bleiben erhalten)
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        # 5️⃣ Tabellenbereich für alle Daten definieren
        last_column_letter = ws.cell(row=1, column=len(df.columns)).column_letter  # Letzte Spalte
        last_row = len(df) + 1  # Anzahl Zeilen + Header
        table_range = f"A1:{last_column_letter}{last_row}"

        # 6️⃣ Tabelle in Excel mit Filter-Dropdowns formatieren
        table = Table(displayName="DatenTabelle", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # 7️⃣ Formatierung für Zellen setzen (Zahlen als Text speichern, um führende Nullen zu erhalten)
        for col in ws.iter_cols(min_row=2, max_row=last_row, min_col=1, max_col=len(df.columns)):
            for cell in col:
                cell.number_format = "@"  # Setzt das Format auf "Text"
                cell.alignment = Alignment(horizontal="left")  # Text linksbündig

        # 8️⃣ **Spaltenbreite automatisch anpassen (ca. doppelt so breit)**
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_length = max(
                [len(str(cell.value)) if cell.value else 0 for cell in ws[col_idx]]
            )
            adjusted_width = (max_length + 2) * 0.5  # **Breite um Faktor ~1.8 erhöhen**
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

        # 9️⃣ Datei speichern
        wb.save(excel_filename)

        print(f"✅ Excel-Datei '{excel_filename}' wurde erfolgreich erstellt!")
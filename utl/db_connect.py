from hdbcli import dbapi
import csv
from datetime import date
from timestamp import add_timestamp_to_filename 

conn = dbapi.connect(
    address="10.0.0.80",
    port=30015,
    user="CUSTOMER_ADMIN",
    password="mIBFtYdgmE4nh0VaNkJS",
)

def execute_query(connection, query, params=None):
    cursor = connection.cursor()
    cursor.execute(query, params or ())
    results = cursor.fetchall()
    #column_names = [desc[0] for desc in cursor.description]
    #print("Spaltennamen:", column_names)
    cursor.close()
    return results



def get_column_names_and_save_to_csv(connection, output_file):
    """
    Ruft die Spaltennamen einer Tabelle aus SAP HANA ab und speichert sie in eine CSV-Datei.
    
    :param connection: SAP HANA Datenbankverbindung (hdbcli dbapi)
    :param table_name: Name der Tabelle, aus der die Spaltennamen geholt werden sollen
    :param output_file: Name der CSV-Datei, in die die Spalten gespeichert werden
    """
    cursor = connection.cursor()
    table_name = '"NEMO"."pa_export"'
    try:
        # Abfrage, um die Spaltennamen zu erhalten
        query = f"SELECT * FROM {table_name} limit 1"
        print(query)
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Spaltennamen extrahieren
        column_names = [desc[0] for desc in cursor.description]

        # CSV-Datei schreiben
        with open(output_file, mode="w", newline="", encoding="utf-8") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["Column Name"])  # Header schreiben
            for col in column_names:
                writer.writerow([f"'{col}',"])  # Jede Spalte in eine eigene Zeile schreiben

        print(f"✅ Spaltennamen wurden erfolgreich in '{output_file}' gespeichert.")
    
    except Exception as e:
        print(f"⚠️ Fehler beim Abrufen der Spaltennamen: {e}")
    
    finally:
        cursor.close()

columns=[
'COMPANY',
'PROCESS_LINKAGE',
"PART_DESC1", 
"PART_DESC2",
"PART_DESC3",
"PART_DESC4",
"PART_I_D",
'PART_OID',
"PART_GROUP",
"PART_GROUP_DESC",
'PART_VARIANT_TYPE',
'PART_VARIANT_TYPE_DESC',
'PART_TYPE',
'PART_TYPE_DESCRIPTION',
'PROCESS',
'PROCESS_DATE',
'ORDER_DOC_I_D',
'ORDER_DOC_LINE_QTY',
'ORDER_DOC_LINE_OPEN',
'ORDER_DOC_LINE_STORAGE_AREA',
'ORDER_DOC_LINE_REQUESTED_DATE',
'ORDER_DOC_LINE_DELIVERY_DATE',
'DELIVERY_DATE',
'REQUESTED_DATE',
'MVMT_OBJECT_I_D',
'MVMT_POSTING_CODE',
'MVMT_POSTING_TYPE',
'MVMT_POSTING_TYPE_DESC',
'MVMT_USAGE',
'MVMT_STORAGE_AREA_ON_HAND',
'MVMT_STORAGE_AREA',
'MVMT_QUANTITY',
'MVMT_POSTING_DATE',
'MVMT_CREATION_DATE_TIME',
'MVMT_MOVE_TYPE',
'PUR_ORDER_DOC_I_D',
'PUR_ORDER_DOC_DATE',
'PUR_ORDER_CREATION_DATE',
'PUR_ORDER_DOC_OPEN',
'PUR_ORDER_LINE_STORAGE_AREA',
'PUR_ORDER_LINE_QTY',
'PROD_ORDER_O_I_D',
'PROD_ORDER_DOC_NO',
'PROD_ORDER_RELEASE_DATE',
'PROD_ORDER_STORAGE_AREA',
'PROD_ORDER_REQ_START_DATE',
'PROD_ORDER_TARGET_QTY',
'PROD_ORDER_FINISHED_QTY',
'PROD_ORDER_SCRAP_QTY',
'PROD_ORDER_CYCLE_TIME',
'PROD_ORDER_COMPL_DATE',
'PROD_ORDER_LINE_CHANGE_DATE',
'PROD_ORDER_LINE_STATUS_DESC',
'PROD_ORDER_LINE_DEMAND_DATE',
'PROD_ORDER_LINE_TARGET_DATE',
'PROD_ORDER_ACT_O_I_D',
'PROD_ORDER_ACT_OPERATION',
'PROD_ORDER_ACT_END_DATE',
'PROD_ORDER_ACT_END_TIME',
'PROD_ORDER_ACT_START_DATE',
'PROD_ORDER_ACT_START_TIME',
'PROD_ORDER_RES_O_I_D',
'PROD_ORDER_RES_RESOURCE',
'PROD_ORDER_PARENT_ORDER_O_I_D',
'PROD_ORDER_PART_ORDER_O_I_D',
'PROD_ORDER_CREATION_DATE'
]
column_list = ", ".join(f'"{col}"' for col in columns)
grenzdatum_start = date(2024, 1, 1)  # Python DATE-Objekt
grenzdatum_ende = date(2024, 2, 1)

query = f"""
    SELECT {column_list} 
    FROM "NEMO"."pa_export_apra" 
    WHERE "COMPANY" = ?
    AND TO_DATE("PROCESS_DATE", 'YYYY-MM-DD') > ?
    AND TO_DATE("PROCESS_DATE", 'YYYY-MM-DD') < ?
    AND (
    PROCESS = 'Make'
    OR PROCESS = 'Production'
    OR PROCESS = 'ProductionOperation'
    OR PROCESS = 'Source'
    OR PROCESS = 'Deliver'
    )
;
 """


"""   
    AND (
    PART_I_D = '31612046'
    OR PART_I_D = '031800112'
    OR PART_I_D = '0240121'
    OR PART_I_D = '044000131'
    OR PART_I_D = '0630015'
    OR PART_I_D = '080205401'
    OR PART_I_D = '0239205'
    )
    ;
"""
"""
31612046
031800112
0240121
044000131
0630015 
080205401
0239205
0241030
031800010
0802009
0803026
080206140
0631678
0239204
0803090
0801006
202074070
080206120
031801212
024599490
0630256
"""
params = ("001", grenzdatum_start.strftime("%Y-%m-%d"), grenzdatum_ende.strftime("%Y-%m-%d"))  # Übergabe als String
results = execute_query(conn, query,params=params)
#results = get_column_names_and_save_to_csv(conn,'Spaltennamen.csv')
        # CSV-Datei schreiben
with open('output_daten.csv', mode="w", newline="", encoding="utf-8") as csv_file:
    writer = csv.writer(csv_file)
    writer.writerow(columns)  # Header schreiben
    for row in results: 
        writer.writerow(row)  # Jede Spalte in eine eigene Zeile schreiben
print('done')

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

# 📂 Dateinamen
csv_filename = "output_daten.csv"
excel_filename = add_timestamp_to_filename("output_daten.xlsx")

# 1️⃣ CSV in Pandas DataFrame laden (string dtype, um führende Nullen zu erhalten)
df = pd.read_csv(csv_filename, dtype=str, encoding="utf-8")

# 2️⃣ Neues Excel-Workbook erstellen
wb = Workbook()
ws = wb.active
ws.title = "Daten"

# 3️⃣ Kopfzeile schreiben (stellt sicher, dass alle Spalten korrekt erfasst werden)
ws.append(df.columns.tolist())

# 4️⃣ Daten einfügen (Führende Nullen bleiben erhalten)
for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# 5️⃣ Tabellenbereich für alle Daten definieren
last_column_letter = ws.cell(row=1, column=len(df.columns)).column_letter  # Letzte Spalte
last_row = len(df) + 1  # Anzahl Zeilen + Header
table_range = f"A1:{last_column_letter}{last_row}"

# 6️⃣ Tabelle in Excel mit Filter-Dropdowns formatieren
table = Table(displayName="DatenTabelle", ref=table_range)
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

# 7️⃣ Formatierung für Zellen setzen (Zahlen als Text speichern, um führende Nullen zu erhalten)
for col in ws.iter_cols(min_row=2, max_row=last_row, min_col=1, max_col=len(df.columns)):
    for cell in col:
        cell.number_format = "@"  # Setzt das Format auf "Text"
        cell.alignment = Alignment(horizontal="left")  # Text linksbündig

# 8️⃣ **Spaltenbreite automatisch anpassen (ca. doppelt so breit)**
for col_idx, col_name in enumerate(df.columns, start=1):
    max_length = max(
        [len(str(cell.value)) if cell.value else 0 for cell in ws[col_idx]]
    )
    adjusted_width = (max_length + 2) * 0.5  # **Breite um Faktor ~1.8 erhöhen**
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

# 9️⃣ Datei speichern
wb.save(excel_filename)

print(f"✅ Excel-Datei '{excel_filename}' wurde erfolgreich erstellt!")



# Benutze Dataframe df zur weiteren verarbeitung

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import norm

def fit_gaussian_distribution(df):
    """
    Berechnet die Differenz in Tagen zwischen 'PROD_ORDER_CREATION_DATE' und 'PROD_ORDER_COMPL_DATE',
    plottet die empirische Dichteverteilung und passt eine Gaußverteilung (Normalverteilung) an.
    
    :param df: pandas DataFrame mit den Spalten 'PROD_ORDER_CREATION_DATE' und 'PROD_ORDER_COMPL_DATE'
    """
    # 📌 Datumswerte in echtes Datetime-Format umwandeln
    df["PROD_ORDER_CREATION_DATE"] = pd.to_datetime(df["PROD_ORDER_CREATION_DATE"], format="%Y-%m-%d", errors="coerce")
    df["PROD_ORDER_COMPL_DATE"] = pd.to_datetime(df["PROD_ORDER_COMPL_DATE"], format="%Y-%m-%d", errors="coerce")

    # 📌 Differenz in Tagen berechnen
    df["DAYS_TO_COMPLETE"] = (df["PROD_ORDER_COMPL_DATE"] - df["PROD_ORDER_CREATION_DATE"]).dt.days

    # 📌 Ungültige Werte (NaN, negative Differenzen) entfernen
    df = df.dropna(subset=["DAYS_TO_COMPLETE"])
    df = df[df["DAYS_TO_COMPLETE"] >= 0]

    # 📊 Histogramm der Verteilung
    plt.figure(figsize=(10, 5))
    counts, bin_edges, _ = plt.hist(df["DAYS_TO_COMPLETE"], bins=20, density=True, alpha=0.6, color="b", label="Empirische Verteilung")

    # 📌 Berechnung der Gaußverteilung
    mu, sigma = norm.fit(df["DAYS_TO_COMPLETE"])  # Berechnet Mittelwert & Standardabweichung
    x = np.linspace(min(df["DAYS_TO_COMPLETE"]), max(df["DAYS_TO_COMPLETE"]), 100)
    pdf = norm.pdf(x, mu, sigma)  # Wahrscheinlichkeitsdichtefunktion (PDF)

    # 📊 Gauß-Fitting hinzufügen
    plt.plot(x, pdf, "r-", label=f"Gauß-Fit: μ={mu:.2f}, σ={sigma:.2f}")

    # 📌 Labels und Titel setzen
    plt.xlabel("Tage bis Abschluss")
    plt.ylabel("Dichte")
    plt.title("Verteilungsfunktion der Produktionsaufträge & Gauß-Fit")
    plt.legend()
    plt.grid()
    plt.show()

    return df, mu, sigma  # Gibt bereinigte Daten + Parameter zurück

fit_gaussian_distribution(df)